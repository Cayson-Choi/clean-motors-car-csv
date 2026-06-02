import csv
import pytest
from openpyxl import Workbook

PARK_HEADER = ['유효','VIP','차량번호','입주사아이디','입주사명','유효일시','만료일시',
               '정기권명','월정액','결제금액','비고','부서명','이름','전화번호','메모','수정일시']
PARK_NOTE = ['(O:추가 X:삭제)','(생략)','(필수)','(생략)','(선택)','(필수)','(필수)',
             '(사용안함)','(사용안함)','(사용안함)','(생략)','(선택)','(선택)','(선택)','(선택)','(생략)']

def _row(plate, name=''):
    r = ['']*16
    r[0]='O'; r[1]='X'; r[2]=plate; r[4]=name
    r[5]='2026-05-01 00:00:00'; r[6]='2999-12-31 00:00:00'
    return r

@pytest.fixture
def park_csv(tmp_path):
    p = tmp_path / 'park.csv'
    rows = [PARK_HEADER, PARK_NOTE, _row('11가1111','오토마트'),
            _row('22나2222','그린모터스'), _row('33다3333','직원용')]
    with open(p, 'w', encoding='utf-8-sig', newline='') as f:
        csv.writer(f).writerows(rows)
    return str(p)

LEDGER_HEADER = ['일련번호','차종','등록번호','소유자','제시신고일','매도신고일','구분','매매업자(사업자)','대장상태']

def _ledger(tmp_path, name, rows):
    wb = Workbook(); ws = wb.active
    ws.append(LEDGER_HEADER)
    for r in rows:
        ws.append(r)
    p = tmp_path / name
    wb.save(p)
    return str(p)

@pytest.fixture
def jesi_xlsx(tmp_path):
    # (일련, 차종, 등록번호, 소유자, 제시일, 매도일, 구분, 매매업자, 대장상태)
    rows = [
        ['1','승용','AA1','소유1','20260501','','상사','프로모터스','제시(상사)'],
        ['2','승용','BB2','소유2','20260502','','상사','그린자동차매매상사','미이전'],   # 미이전+공란 -> 추가
        ['3','승용','CC3','소유3','20260503','20260510','상사','장성대형화물','미이전'],  # 미이전+매도일있음 -> 제외
        ['4','승용','DD4','소유4','20260504','20260512','상사','프로모터스','매도'],       # 매도 -> 제외
    ]
    return _ledger(tmp_path, 'jesi.xlsx', rows)

@pytest.fixture
def outcha_a(tmp_path):
    # 차량번호가 4번째 열, 2행은 빈 행
    wb = Workbook(); ws = wb.active
    ws.append(['순번','입고일','출고일','차량번호','차량명','진행기관'])
    ws.append([None,None,None,None,None,None])
    ws.append(['1','2026-01-01','2026-01-05','22나2222','차A','기관'])
    ws.append(['2','2026-01-02','2026-01-06','99바9999','차B','기관'])
    p = tmp_path/'outcha_a.xlsx'; wb.save(p); return str(p)

@pytest.fixture
def outcha_b(tmp_path):
    # 차량번호가 2번째 열
    wb = Workbook(); ws = wb.active
    ws.append(['상태','차량번호','차량명','진행기관','진행','입고일','출고일'])
    ws.append(['출고완료','77사7777','차C','기관','낙찰','2026-02-01','2026-02-05'])
    p = tmp_path/'outcha_b.xlsx'; wb.save(p); return str(p)
