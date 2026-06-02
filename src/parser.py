import csv
from openpyxl import load_workbook

def detect_encoding(path):
    for enc in ('utf-8-sig', 'utf-8', 'cp949'):
        try:
            with open(path, encoding=enc) as f:
                f.read()
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return 'utf-8'

def parse_parking_csv(path):
    enc = detect_encoding(path)
    with open(path, encoding=enc, newline='') as f:
        rows = [list(r) for r in csv.reader(f)]
    header = rows[:2]
    data = [r for r in rows[2:] if len(r) > 2 and str(r[2]).strip()]
    # 길이 16으로 정규화
    for r in data:
        while len(r) < 16:
            r.append('')
    return header, data, enc

def _norm(v):
    return '' if v is None else str(v).strip()

def parse_ledger(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    head = [_norm(c) for c in rows[0]]
    def col(r, name):
        if name in head:
            i = head.index(name)
            return r[i] if i < len(r) else None
        return None
    out = []
    for r in rows[1:]:
        if not any(c is not None and _norm(c) for c in r):
            continue
        plate = _norm(col(r, '등록번호'))
        if not plate:
            continue
        out.append({
            'plate': plate,
            'sold_date': _norm(col(r, '매도신고일')),
            'dealer': _norm(col(r, '매매업자(사업자)')),
            'status': _norm(col(r, '대장상태')),
        })
    return out

def extract_presented_additions(ledger_rows):
    add = {}
    for row in ledger_rows:
        st = row['status']
        sold_empty = row['sold_date'] == ''
        if st == '제시(상사)' or (st == '미이전' and sold_empty):
            add.setdefault(row['plate'], row['dealer'])
    return add

def ledger_all_plates(ledger_rows):
    return set(r['plate'] for r in ledger_rows)
